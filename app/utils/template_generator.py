import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "templates", "excel")
JSON_TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "templates", "json")

HEADER_FILL = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
SAMPLE_FONT = Font(name="Arial", size=10)
BORDER_THIN = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def style_worksheet(ws, headers, rows):
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_data in rows:
        ws.append(row_data)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = SAMPLE_FONT
            cell.border = BORDER_THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Auto-adjust column width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            first_line = val_str.split("\n")[0]
            max_len = max(max_len, len(first_line))
        ws.column_dimensions[col_letter].width = max(min(max_len + 5, 45), 14)


def get_sample_vocabulary_data():
    headers = [
        "word", "pronunciation", "part_of_speech", "meaning_vi",
        "example_en", "example_vi", "topic", "level",
        "image_url", "collocations", "synonyms", "antonyms"
    ]
    rows = [
        [
            "accomplish", "/əˈkʌm.plɪʃ/", "verb", "hoàn thành, đạt được",
            "She accomplished all her goals for this semester.", "Cô ấy đã hoàn thành mọi mục tiêu trong học kỳ này.",
            "Education", "B1", "https://images.unsplash.com/photo-1434030216411-0b793f4b4173?w=500",
            "accomplish a mission; accomplish a task", "achieve, complete, fulfill", "fail, abandon"
        ],
        [
            "sustainable", "/səˈsteɪ.nə.bəl/", "adjective", "bền vững, thân thiện với môi trường",
            "The company is committed to sustainable development.", "Công ty cam kết phát triển bền vững.",
            "Environment", "B2", "https://images.unsplash.com/photo-1542601906990-b4d3fb778b09?w=500",
            "sustainable energy; sustainable development", "eco-friendly, renewable", "unsustainable, wasteful"
        ],
        [
            "negotiate", "/nəˈɡoʊ.ʃi.eɪt/", "verb", "đàm phán, thương lượng",
            "The union is negotiating for better working conditions.", "Công đoàn đang đàm phán để có điều kiện làm việc tốt hơn.",
            "Business", "B2", "",
            "negotiate a contract; negotiate terms", "bargain, discuss", "surrender"
        ],
        [
            "pivotal", "/ˈpɪv.ə.t̬əl/", "adjective", "then chốt, mang tính quyết định",
            "He played a pivotal role in the success of the project.", "Anh ấy đóng một vai trò then chốt trong sự thành công của dự án.",
            "General", "C1", "",
            "pivotal role; pivotal moment", "crucial, vital, critical", "trivial, insignificant"
        ]
    ]
    return headers, rows


def get_sample_grammar_data():
    headers = [
        "title", "category", "level", "difficulty", "summary",
        "rule_explanation", "examples_json", "common_mistakes", "tips_tricks"
    ]
    rows = [
        [
            "Thì Hiện Tại Hoàn Thành (Present Perfect Tense)",
            "Các thì (Tenses)",
            "A2",
            "Medium",
            "Diễn tả hành động đã xảy ra trong quá khứ nhưng có kết quả hoặc liên quan mật thiết đến hiện tại.",
            "1. CÔNG THỨC KHẲNG ĐỊNH:\nS + have / has + V3/ed + (O)\n\n2. CÔNG THỨC PHỦ ĐỊNH:\nS + have / has + not (haven't / hasn't) + V3/ed\n\n3. CÂU HỎI:\nHave / Has + S + V3/ed...?",
            "I have lived in Hanoi for 5 years.|Tôi đã sống ở Hà Nội được 5 năm.\nShe has already finished her homework.|Cô ấy đã làm xong bài tập về nhà rồi.",
            "Nhầm lẫn giữa Thì Quá Khứ Đơn (Past Simple - có thời gian xác định cụ thể trong quá khứ như yesterday, in 2020) và Hiện Tại Hoàn Thành.",
            "Dấu hiệu nhận biết đặc trưng: since, for, already, yet, just, ever, never, so far, up to now."
        ],
        [
            "Câu Điều Kiện Loại 2 (Conditional Type 2)",
            "Câu điều kiện (Conditionals)",
            "B1",
            "Medium",
            "Diễn tả giả định trái ngược với thực tế ở hiện tại hoặc khó có thể xảy ra.",
            "Mệnh đề If: If + S + V2/ed (động từ to be dùng 'were' cho mọi ngôi)\nMệnh đề chính: S + would / could + V_inf",
            "If I were you, I would accept the job offer.|Nếu tôi là bạn, tôi sẽ nhận lời đề nghị làm việc đó.\nIf he had more money, he would travel around the world.|Nếu anh ấy có nhiều tiền hơn, anh ấy sẽ đi du lịch vòng quanh thế giới.",
            "Dùng 'was' thay vì 'were' trong các bài thi trang trọng (Formal English quy định dùng 'were').",
            "Hãy nhớ câu cửa miệng: 'If I were you, I would...' để luôn nhớ chia 'were' và 'would + V_inf'."
        ]
    ]
    return headers, rows


def get_sample_lessons_data():
    headers = [
        "title", "level", "skill", "short_description",
        "content", "examples", "thumbnail_url"
    ]
    rows = [
        [
            "Mastering Business Email Communication",
            "B2",
            "Writing",
            "Học cách viết email giao dịch thương mại chuyên nghiệp, trang trọng và thuyết phục.",
            "### 1. Introduction\nWriting effective business emails is essential in professional settings.\n\n### 2. Email Structure\n- **Subject Line**: Concise and clear\n- **Salutation**: Dear Mr./Ms. [Last Name] or Dear [First Name]\n- **Opening**: I hope this email finds you well.\n- **Main Body**: State the purpose clearly.\n- **Call to Action**: Please let me know your availability.\n- **Sign-off**: Best regards / Sincerely.",
            "Could you please confirm receipt of this document?|Bạn có thể vui lòng xác nhận đã nhận tài liệu này không?\nI look forward to hearing from you soon.|Tôi rất mong sớm nhận được phản hồi từ bạn.",
            "https://images.unsplash.com/photo-1516321318423-f06f85e504b3?w=500"
        ],
        [
            "Everyday Small Talk & Networking",
            "A2",
            "Speaking",
            "Nắm vững các mẫu câu bắt chuyện tự nhiên trong môi trường công sở và đời sống.",
            "### 1. Weather and Surroundings\nStart conversations with neutral topics like weather or current surroundings.\n\n### 2. Open-ended Questions\nUse 'How was your weekend?' instead of 'Did you have a good weekend?' to keep conversations flowing.",
            "Nice weather we're having today, isn't it?|Thời tiết hôm nay đẹp thật, phải không?\nHow are things going with your new project?|Dự án mới của bạn tiến triển thế nào rồi?",
            "https://images.unsplash.com/photo-1522202176988-66273c2fd55f?w=500"
        ]
    ]
    return headers, rows


def get_sample_questions_data():
    headers = [
        "question_text", "option_a", "option_b", "option_c", "option_d",
        "correct_option", "explanation", "topic", "level"
    ]
    rows = [
        [
            "She _______ in this company since 2018.",
            "works",
            "worked",
            "has worked",
            "is working",
            "C",
            "Dấu hiệu nhận biết 'since 2018' chỉ hành động bắt đầu từ quá khứ và kéo dài đến hiện tại -> chia Hiện tại hoàn thành (has worked).",
            "Tenses",
            "A2"
        ],
        [
            "If the weather _______ fine tomorrow, we will go for a picnic.",
            "is",
            "was",
            "will be",
            "were",
            "A",
            "Câu điều kiện loại 1 (Conditional Type 1) diễn tả khả năng có thật ở tương lai: Mệnh đề If chia Hiện tại đơn (is).",
            "Conditionals",
            "A2"
        ],
        [
            "The committee reached a _______ decision after several hours of intense debate.",
            "unanimous",
            "reluctant",
            "hesitant",
            "ambiguous",
            "A",
            "'Unanimous' nghĩa là nhất trí/đồng thuận hoàn toàn. Câu diễn tả ủy ban đã đạt được quyết định nhất trí sau nhiều giờ tranh luận.",
            "Vocabulary",
            "B2"
        ]
    ]
    return headers, rows


def get_sample_exams_data():
    headers = [
        "category", "title", "duration_minutes", "difficulty",
        "skill", "part", "question_text", "option_a", "option_b", "option_c", "option_d",
        "correct_answer", "explanation"
    ]
    rows = [
        [
            "TOEIC",
            "TOEIC Reading Mini Test 01",
            "20",
            "Medium",
            "READING",
            "Part 5",
            "All employees are requested to submit their expense reports _______ Friday afternoon.",
            "by",
            "until",
            "at",
            "on",
            "A",
            "'By Friday afternoon' nghĩa là trước hoặc muộn nhất vào chiều thứ Sáu (deadline). 'Until' chỉ hành động kéo dài liên tục.",
        ],
        [
            "TOEIC",
            "TOEIC Reading Mini Test 01",
            "20",
            "Medium",
            "READING",
            "Part 5",
            "Mr. Henderson was _______ promoted to Senior Marketing Director.",
            "recent",
            "recently",
            "more recent",
            "recency",
            "B",
            "Vị trí đứng giữa trợ động từ 'was' và phân từ hai 'promoted' cần một trạng từ (Adverb) để bổ nghĩa -> chọn 'recently'.",
        ],
        [
            "IELTS",
            "IELTS General Practice Test 01",
            "60",
            "Hard",
            "READING",
            "Section 1",
            "According to the notice, what must visitors do before entering the construction area?",
            "Sign the guestbook",
            "Wear a hard hat and safety vest",
            "Call their supervisor",
            "Leave their mobile phones at reception",
            "B",
            "Đoạn văn quy định rõ tất cả khách tham quan bắt buộc phải trang bị mũ bảo hộ và áo phản quang trước khi bước vào công trường.",
        ]
    ]
    return headers, rows


def generate_all_templates():
    os.makedirs(TEMPLATE_DIR, exist_ok=True)
    os.makedirs(JSON_TEMPLATE_DIR, exist_ok=True)

    data_map = {
        "vocabulary": (get_sample_vocabulary_data, "template_vocabulary"),
        "grammar": (get_sample_grammar_data, "template_grammar"),
        "lessons": (get_sample_lessons_data, "template_lessons"),
        "questions": (get_sample_questions_data, "template_questions"),
        "exams": (get_sample_exams_data, "template_exams"),
    }

    created_files = []

    for key, (data_func, base_name) in data_map.items():
        headers, rows = data_func()

        # 1. Generate Excel (.xlsx)
        wb = Workbook()
        ws = wb.active
        ws.title = key.capitalize()
        style_worksheet(ws, headers, rows)
        xlsx_path = os.path.join(TEMPLATE_DIR, f"{base_name}.xlsx")
        wb.save(xlsx_path)
        created_files.append(xlsx_path)

        # 2. Generate JSON (.json)
        json_records = []
        for r in rows:
            record = {}
            for col_idx, h in enumerate(headers):
                val = r[col_idx] if col_idx < len(r) else ""
                record[h] = val
            json_records.append(record)

        json_path = os.path.join(JSON_TEMPLATE_DIR, f"{base_name}.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_records, f, ensure_ascii=False, indent=2)
        created_files.append(json_path)

    return created_files


if __name__ == "__main__":
    files = generate_all_templates()
    print(f"Generated {len(files)} template files in {TEMPLATE_DIR} and {JSON_TEMPLATE_DIR}")

