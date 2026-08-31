import io
import json
import uuid
from datetime import datetime
from openpyxl import load_workbook

from app.extensions import db
from app.modules.learning.models import Vocabulary, GrammarTopic, Lesson, Question
from app.modules.exams.models import Exam, ExamQuestion
from app.modules.admin.models import AuditLog


CONTENT_SCHEMAS = {
    "vocabulary": {
        "title": "Từ vựng (Vocabulary)",
        "required_columns": ["word", "pronunciation", "part_of_speech", "meaning_vi", "example_en", "example_vi", "topic", "level"],
        "optional_columns": ["image_url", "collocations", "synonyms", "antonyms"],
        "level_valid": ["A1", "A2", "B1", "B2", "C1", "C2"],
    },
    "grammar": {
        "title": "Ngữ pháp (Grammar Topics)",
        "required_columns": ["title", "category", "level", "difficulty", "summary", "rule_explanation", "examples_json"],
        "optional_columns": ["common_mistakes", "tips_tricks"],
        "level_valid": ["A1", "A2", "B1", "B2", "C1", "C2"],
        "difficulty_valid": ["Easy", "Medium", "Hard"],
    },
    "lessons": {
        "title": "Bài học (Lessons)",
        "required_columns": ["title", "level", "skill", "short_description", "content", "examples"],
        "optional_columns": ["thumbnail_url"],
        "level_valid": ["A1", "A2", "B1", "B2", "C1", "C2"],
        "skill_valid": ["Grammar", "Vocabulary", "Reading", "Listening", "Speaking", "Writing", "General"],
    },
    "questions": {
        "title": "Câu hỏi Trắc nghiệm (Questions)",
        "required_columns": ["question_text", "option_a", "option_b", "option_c", "option_d", "correct_option", "explanation", "topic", "level"],
        "optional_columns": ["skill"],
        "level_valid": ["A1", "A2", "B1", "B2", "C1", "C2"],
        "correct_option_valid": ["A", "B", "C", "D"],
    },
    "exams": {
        "title": "Đề thi & Kiểm tra (Exams)",
        "required_columns": ["category", "title", "duration_minutes", "difficulty", "skill", "part", "question_text", "option_a", "option_b", "option_c", "option_d", "correct_answer", "explanation"],
        "optional_columns": ["transcript", "media_url"],
        "difficulty_valid": ["Easy", "Medium", "Hard"],
    }
}


def parse_and_validate_file(file_stream, filename, content_type):
    if filename.lower().endswith(".json"):
        return parse_and_validate_json(file_stream, content_type)
    return parse_and_validate_excel(file_stream, content_type)


def parse_and_validate_json(file_stream, content_type):
    if content_type not in CONTENT_SCHEMAS:
        return {
            "success": False,
            "error": f"Loại nội dung '{content_type}' không được hỗ trợ."
        }

    schema = CONTENT_SCHEMAS[content_type]
    try:
        raw_content = file_stream.read()
        if isinstance(raw_content, bytes):
            raw_content = raw_content.decode("utf-8")
        data = json.loads(raw_content)
    except Exception as e:
        return {
            "success": False,
            "error": f"Không thể đọc file JSON. Vui lòng đảm bảo định dạng file là .json hợp lệ. ({str(e)})"
        }

    if isinstance(data, dict):
        for k in ["items", "data", "records", "vocabulary", "grammar", "lessons", "questions", "exams"]:
            if k in data and isinstance(data[k], list):
                data = data[k]
                break
        else:
            data = [data]

    if not isinstance(data, list) or len(data) == 0:
        return {
            "success": False,
            "error": "File JSON trống hoặc không chứa danh sách các bản ghi (Array of Objects)."
        }

    valid_records = []
    error_records = []

    for idx, item in enumerate(data, start=1):
        if not isinstance(item, dict):
            error_records.append({
                "row_number": idx,
                "data": {"raw": str(item)},
                "errors": ["Bản ghi không phải là JSON Object (dict)"]
            })
            continue

        row_data = {}
        normalized_item = {str(k).strip().lower(): str(v).strip() if v is not None else "" for k, v in item.items()}

        for col_name in schema["required_columns"] + schema.get("optional_columns", []):
            row_data[col_name] = normalized_item.get(col_name, "")

        row_errors = []

        for req_col in schema["required_columns"]:
            if not row_data.get(req_col):
                row_errors.append(f"Trường '{req_col}' không được để trống")

        if "level" in row_data and row_data["level"]:
            level_upper = row_data["level"].upper()
            if "level_valid" in schema and level_upper not in schema["level_valid"]:
                row_errors.append(f"Cấp độ '{row_data['level']}' không hợp lệ (Phải là A1, A2, B1, B2, C1, hoặc C2)")
            row_data["level"] = level_upper

        if "difficulty" in row_data and row_data["difficulty"]:
            diff_title = row_data["difficulty"].title()
            if "difficulty_valid" in schema and diff_title not in schema["difficulty_valid"]:
                row_errors.append(f"Độ khó '{row_data['difficulty']}' không hợp lệ (Phải là Easy, Medium, hoặc Hard)")
            row_data["difficulty"] = diff_title

        if "skill" in row_data and row_data["skill"]:
            skill_title = row_data["skill"].title()
            if "skill_valid" in schema and skill_title not in schema["skill_valid"]:
                row_errors.append(f"Kỹ năng '{row_data['skill']}' không hợp lệ (Phải là Grammar, Vocabulary, Reading, Listening, Speaking, Writing, General)")
            row_data["skill"] = skill_title

        if "correct_option" in row_data and row_data["correct_option"]:
            opt_upper = row_data["correct_option"].upper()
            if "correct_option_valid" in schema and opt_upper not in schema["correct_option_valid"]:
                row_errors.append(f"Đáp án đúng '{row_data['correct_option']}' không hợp lệ (Phải là A, B, C, hoặc D)")
            row_data["correct_option"] = opt_upper

        if row_errors:
            error_records.append({
                "row_number": idx,
                "data": row_data,
                "errors": row_errors
            })
        else:
            valid_records.append({
                "row_number": idx,
                "data": row_data
            })

    batch_id = str(uuid.uuid4())

    return {
        "success": True,
        "batch_id": batch_id,
        "content_type": content_type,
        "content_title": schema["title"],
        "total_rows": len(valid_records) + len(error_records),
        "valid_count": len(valid_records),
        "error_count": len(error_records),
        "valid_records": valid_records,
        "error_records": error_records,
        "preview_sample": [r["data"] for r in valid_records[:10]]
    }


def parse_and_validate_excel(file_stream, content_type):
    if content_type not in CONTENT_SCHEMAS:
        return {
            "success": False,
            "error": f"Loại nội dung '{content_type}' không được hỗ trợ."
        }

    schema = CONTENT_SCHEMAS[content_type]
    try:
        wb = load_workbook(filename=io.BytesIO(file_stream.read()), data_only=True)
    except Exception as e:
        return {
            "success": False,
            "error": f"Không thể đọc file Excel. Vui lòng đảm bảo định dạng file là .xlsx hợp lệ. ({str(e)})"
        }

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return {
            "success": False,
            "error": "File Excel trống, không có dữ liệu."
        }

    # Header check
    raw_headers = [str(h or "").strip().lower() for h in rows[0]]
    missing_cols = [col for col in schema["required_columns"] if col not in raw_headers]
    if missing_cols:
        return {
            "success": False,
            "error": f"File Excel thiếu các cột bắt buộc: {', '.join(missing_cols)}."
        }

    header_indices = {col: raw_headers.index(col) for col in raw_headers if col}

    valid_records = []
    error_records = []

    # Iterate data rows
    for row_idx, row in enumerate(rows[1:], start=2):
        # Check if entire row is empty
        if not any(row):
            continue

        row_data = {}
        for col_name, col_idx in header_indices.items():
            val = row[col_idx] if col_idx < len(row) else ""
            if val is not None:
                val = str(val).strip()
            else:
                val = ""
            row_data[col_name] = val

        row_errors = []

        # Validate required columns
        for req_col in schema["required_columns"]:
            if not row_data.get(req_col):
                row_errors.append(f"Cột '{req_col}' không được để trống")

        # Validate Level
        if "level" in row_data and row_data["level"]:
            level_upper = row_data["level"].upper()
            if "level_valid" in schema and level_upper not in schema["level_valid"]:
                row_errors.append(f"Cấp độ '{row_data['level']}' không hợp lệ (Phải là A1, A2, B1, B2, C1, hoặc C2)")
            row_data["level"] = level_upper

        # Validate Difficulty
        if "difficulty" in row_data and row_data["difficulty"]:
            diff_title = row_data["difficulty"].title()
            if "difficulty_valid" in schema and diff_title not in schema["difficulty_valid"]:
                row_errors.append(f"Độ khó '{row_data['difficulty']}' không hợp lệ (Phải là Easy, Medium, hoặc Hard)")
            row_data["difficulty"] = diff_title

        # Validate Skill for lessons
        if "skill" in row_data and row_data["skill"]:
            skill_title = row_data["skill"].title()
            if "skill_valid" in schema and skill_title not in schema["skill_valid"]:
                row_errors.append(f"Kỹ năng '{row_data['skill']}' không hợp lệ (Phải là Grammar, Vocabulary, Reading, Listening, Speaking, Writing, General)")
            row_data["skill"] = skill_title

        # Validate Correct Option
        if "correct_option" in row_data and row_data["correct_option"]:
            opt_upper = row_data["correct_option"].upper()
            if "correct_option_valid" in schema and opt_upper not in schema["correct_option_valid"]:
                row_errors.append(f"Đáp án đúng '{row_data['correct_option']}' không hợp lệ (Phải là A, B, C, hoặc D)")
            row_data["correct_option"] = opt_upper

        if row_errors:
            error_records.append({
                "row_number": row_idx,
                "data": row_data,
                "errors": row_errors
            })
        else:
            valid_records.append({
                "row_number": row_idx,
                "data": row_data
            })

    batch_id = str(uuid.uuid4())

    return {
        "success": True,
        "batch_id": batch_id,
        "content_type": content_type,
        "content_title": schema["title"],
        "total_rows": len(valid_records) + len(error_records),
        "valid_count": len(valid_records),
        "error_count": len(error_records),
        "valid_records": valid_records,
        "error_records": error_records,
        "preview_sample": [r["data"] for r in valid_records[:10]]
    }


def commit_import_records(content_type, valid_records, user_id=None, mode="insert_or_update"):
    """
    Persist validated records into PostgreSQL/SQLite database.
    """
    inserted_count = 0
    updated_count = 0

    if content_type == "vocabulary":
        for rec in valid_records:
            d = rec["data"]
            word_str = d["word"].strip()
            existing = Vocabulary.query.filter(Vocabulary.word.ilike(word_str)).first()

            if existing and mode == "insert_or_update":
                existing.pronunciation = d.get("pronunciation", existing.pronunciation)
                existing.part_of_speech = d.get("part_of_speech", existing.part_of_speech)
                existing.meaning_vi = d.get("meaning_vi", existing.meaning_vi)
                existing.example_en = d.get("example_en", existing.example_en)
                existing.example_vi = d.get("example_vi", existing.example_vi)
                existing.topic = d.get("topic", existing.topic)
                existing.level = d.get("level", existing.level)
                if d.get("image_url"):
                    existing.image_url = d["image_url"]
                if d.get("collocations"):
                    existing.collocations = d["collocations"]
                if d.get("synonyms"):
                    existing.synonyms = d["synonyms"]
                if d.get("antonyms"):
                    existing.antonyms = d["antonyms"]
                updated_count += 1
            elif not existing:
                item = Vocabulary(
                    word=word_str,
                    pronunciation=d.get("pronunciation", ""),
                    part_of_speech=d.get("part_of_speech", "noun"),
                    meaning_vi=d.get("meaning_vi", ""),
                    example_en=d.get("example_en", ""),
                    example_vi=d.get("example_vi", ""),
                    topic=d.get("topic", "General"),
                    level=d.get("level", "A1"),
                    image_url=d.get("image_url") or None,
                    collocations=d.get("collocations") or None,
                    synonyms=d.get("synonyms") or None,
                    antonyms=d.get("antonyms") or None,
                )
                db.session.add(item)
                inserted_count += 1

    elif content_type == "grammar":
        for rec in valid_records:
            d = rec["data"]
            title_str = d["title"].strip()
            existing = GrammarTopic.query.filter_by(title=title_str).first()

            if existing and mode == "insert_or_update":
                existing.category = d.get("category", existing.category)
                existing.level = d.get("level", existing.level)
                existing.difficulty = d.get("difficulty", existing.difficulty)
                existing.summary = d.get("summary", existing.summary)
                existing.rule_explanation = d.get("rule_explanation", existing.rule_explanation)
                existing.examples_json = d.get("examples_json", existing.examples_json)
                if d.get("common_mistakes"):
                    existing.common_mistakes = d["common_mistakes"]
                if d.get("tips_tricks"):
                    existing.tips_tricks = d["tips_tricks"]
                updated_count += 1
            elif not existing:
                item = GrammarTopic(
                    title=title_str,
                    category=d.get("category", "General"),
                    level=d.get("level", "A1"),
                    difficulty=d.get("difficulty", "Medium"),
                    summary=d.get("summary", ""),
                    rule_explanation=d.get("rule_explanation", ""),
                    examples_json=d.get("examples_json", ""),
                    common_mistakes=d.get("common_mistakes") or None,
                    tips_tricks=d.get("tips_tricks") or None,
                    is_active=True,
                )
                db.session.add(item)
                inserted_count += 1

    elif content_type == "lessons":
        for rec in valid_records:
            d = rec["data"]
            title_str = d["title"].strip()
            existing = Lesson.query.filter_by(title=title_str).first()

            if existing and mode == "insert_or_update":
                existing.level = d.get("level", existing.level)
                existing.skill = d.get("skill", existing.skill)
                existing.short_description = d.get("short_description", existing.short_description)
                existing.content = d.get("content", existing.content)
                existing.examples = d.get("examples", existing.examples)
                if d.get("thumbnail_url"):
                    existing.thumbnail_url = d["thumbnail_url"]
                updated_count += 1
            elif not existing:
                item = Lesson(
                    title=title_str,
                    level=d.get("level", "A1"),
                    skill=d.get("skill", "General"),
                    short_description=d.get("short_description", ""),
                    content=d.get("content", ""),
                    examples=d.get("examples", ""),
                    thumbnail_url=d.get("thumbnail_url") or None,
                    is_active=True,
                )
                db.session.add(item)
                inserted_count += 1

    elif content_type == "questions":
        for rec in valid_records:
            d = rec["data"]
            q_text = d["question_text"].strip()
            existing = Question.query.filter_by(question_text=q_text).first()

            if existing and mode == "insert_or_update":
                existing.option_a = d.get("option_a", existing.option_a)
                existing.option_b = d.get("option_b", existing.option_b)
                existing.option_c = d.get("option_c", existing.option_c)
                existing.option_d = d.get("option_d", existing.option_d)
                existing.correct_option = d.get("correct_option", existing.correct_option)
                existing.explanation = d.get("explanation", existing.explanation)
                existing.topic = d.get("topic", existing.topic)
                existing.level = d.get("level", existing.level)
                updated_count += 1
            elif not existing:
                item = Question(
                    question_text=q_text,
                    option_a=d.get("option_a", ""),
                    option_b=d.get("option_b", ""),
                    option_c=d.get("option_c", ""),
                    option_d=d.get("option_d", ""),
                    correct_option=d.get("correct_option", "A"),
                    explanation=d.get("explanation", ""),
                    topic=d.get("topic", "General"),
                    level=d.get("level", "A1"),
                )
                db.session.add(item)
                inserted_count += 1

    elif content_type == "exams":
        # Group questions by exam title
        exam_groups = {}
        for rec in valid_records:
            d = rec["data"]
            exam_title = d["title"].strip()
            if exam_title not in exam_groups:
                exam_groups[exam_title] = {
                    "category": d.get("category", "General"),
                    "duration_minutes": int(d.get("duration_minutes", 15)),
                    "difficulty": d.get("difficulty", "Medium"),
                    "questions": []
                }
            exam_groups[exam_title]["questions"].append(d)

        for title_str, grp in exam_groups.items():
            exam = Exam.query.filter_by(title=title_str).first()
            if not exam:
                exam = Exam(
                    title=title_str,
                    category=grp["category"],
                    duration=grp["duration_minutes"],
                    duration_minutes=grp["duration_minutes"],
                    difficulty=grp["difficulty"],
                    question_count=len(grp["questions"]),
                    is_published=True,
                    is_active=True
                )
                db.session.add(exam)
                db.session.flush()
                inserted_count += 1

            for q_data in grp["questions"]:
                eq = ExamQuestion(
                    exam_id=exam.id,
                    skill=q_data.get("skill", "READING"),
                    part=q_data.get("part", "Part 1"),
                    question_text=q_data.get("question_text", ""),
                    option_a=q_data.get("option_a", ""),
                    option_b=q_data.get("option_b", ""),
                    option_c=q_data.get("option_c", ""),
                    option_d=q_data.get("option_d", ""),
                    correct_answer=q_data.get("correct_answer", "A"),
                    explanation=q_data.get("explanation", "")
                )
                db.session.add(eq)

    # Log audit
    if user_id:
        log = AuditLog(
            user_id=user_id,
            action="IMPORT_EXCEL",
            target_type=content_type,
            details=f"Imported {inserted_count} new and updated {updated_count} records via Excel."
        )
        db.session.add(log)

    db.session.commit()

    return {
        "success": True,
        "inserted_count": inserted_count,
        "updated_count": updated_count,
        "total_processed": inserted_count + updated_count
    }
